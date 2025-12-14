
from __future__ import annotations
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

CLASS_RE = re.compile(r"\b(\d{1,2})\s*-\s*([A-ZА-Я])\s*([0-9]{0,2})\b", re.IGNORECASE)

def _norm_name(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    # remove common suffixes
    s = re.sub(r"\b(o['’`]?g['’`]?li|qizi)\b", "", s, flags=re.IGNORECASE).strip()
    return s.lower()

def shorten_teacher_name(full: str) -> str:
    # Expect: Surname Name Patronymic ...
    parts = re.sub(r"\s+", " ", (full or "").strip()).split(" ")
    if len(parts) < 2:
        return full.strip()
    surname = parts[0]
    name = parts[1]
    patronymic = parts[2] if len(parts) >= 3 else ""
    ini1 = (name[0] + ".") if name else ""
    ini2 = (patronymic[0] + ".") if patronymic else ""
    return f"{surname} {ini1}{ini2}".replace("..", ".")

def detect_class_from_workbook(wb: Workbook) -> Optional[str]:
    # Try sheet names first
    for sn in wb.sheetnames:
        m = CLASS_RE.search(sn.replace("_", "-"))
        if m:
            return f"{int(m.group(1))}-{m.group(2).upper()}{m.group(3)}".replace(" ", "")
    # Try scanning first 10x10 of first sheet
    ws = wb[wb.sheetnames[0]]
    for r in range(1, 11):
        for c in range(1, 11):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                m = CLASS_RE.search(v.replace("_", "-"))
                if m:
                    return f"{int(m.group(1))}-{m.group(2).upper()}{m.group(3)}".replace(" ", "")
    return None

def _find_header_row(ws, max_rows=30, max_cols=30) -> int:
    # Heuristic: row with most text headers
    best_row, best_score = 1, -1
    for r in range(1, max_rows + 1):
        score = 0
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and len(v.strip()) >= 2:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row

def _detect_columns(ws) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    header_row = _find_header_row(ws)
    name_col = bsb_col = chsb_col = None
    for c in range(1, 60):
        v = ws.cell(header_row, c).value
        if not isinstance(v, str):
            continue
        hv = v.strip().lower()
        if name_col is None and any(k in hv for k in ["ism", "fio", "f.i.o", "ф.и.о", "name", "familiya"]):
            name_col = c
        if bsb_col is None and any(k in hv for k in ["bsb", "бсб"]):
            bsb_col = c
        if chsb_col is None and any(k in hv for k in ["chsb", "чсб", "chsб"]):
            chsb_col = c
    # Fallback: if name_col missing, try first column with many strings
    if name_col is None:
        best = (None, -1)
        for c in range(1, 15):
            cnt = 0
            for r in range(header_row + 1, header_row + 60):
                v = ws.cell(r, c).value
                if isinstance(v, str) and len(v.strip()) >= 3:
                    cnt += 1
            if cnt > best[1]:
                best = (c, cnt)
        if best[1] >= 3:
            name_col = best[0]
    return name_col, bsb_col, chsb_col

def extract_students_scores(data_wb: Workbook) -> Tuple[str, Dict[str, Tuple[Optional[float], Optional[float]]]]:
    cls = detect_class_from_workbook(data_wb) or ""
    ws = data_wb[data_wb.sheetnames[0]]
    header_row = _find_header_row(ws)
    name_col, bsb_col, chsb_col = _detect_columns(ws)

    results: Dict[str, Tuple[Optional[float], Optional[float]]] = {}
    for r in range(header_row + 1, header_row + 300):
        name_v = ws.cell(r, name_col).value if name_col else None
        if not isinstance(name_v, str):
            continue
        nm = _norm_name(name_v)
        if not nm:
            continue
        bsb_v = ws.cell(r, bsb_col).value if bsb_col else None
        chsb_v = ws.cell(r, chsb_col).value if chsb_col else None
        def _to_num(x):
            if x is None:
                return None
            if isinstance(x, (int, float)):
                return float(x)
            if isinstance(x, str):
                x = x.strip().replace(",", ".")
                try:
                    return float(x)
                except Exception:
                    return None
            return None
        results[nm] = (_to_num(bsb_v), _to_num(chsb_v))
    return cls, results

def _find_teacher_cells(ws):
    keys = ["o‘qituvchi", "oqituvchi", "учитель", "teacher"]
    found = []
    for r in range(1, 40):
        for c in range(1, 20):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                lv = v.lower()
                if any(k in lv for k in keys):
                    found.append((r, c))
    return found

def _find_class_cells(ws):
    keys = ["sinf", "класс", "class"]
    found = []
    for r in range(1, 40):
        for c in range(1, 20):
            v = ws.cell(r, c).value
            if isinstance(v, str) and any(k in v.lower() for k in keys):
                found.append((r, c))
    return found

def _set_adjacent(ws, rc_list, new_value: str):
    # Prefer right cell, else same cell
    for (r, c) in rc_list:
        right = ws.cell(r, c+1)
        if right.value is None or isinstance(right.value, str) or isinstance(right.value, (int, float)):
            right.value = new_value
        else:
            ws.cell(r, c).value = new_value

def _fill_sheet_from_scores(template_ws, scores: Dict[str, Tuple[Optional[float], Optional[float]]]):
    header_row = _find_header_row(template_ws)
    name_col, bsb_col, chsb_col = _detect_columns(template_ws)

    if name_col is None:
        return {"filled": 0, "skipped": 0}

    filled = skipped = 0
    for r in range(header_row + 1, header_row + 600):
        v = template_ws.cell(r, name_col).value
        if not isinstance(v, str):
            continue
        nm = _norm_name(v)
        if nm in scores:
            bsb, chsb = scores[nm]
            if bsb_col:
                template_ws.cell(r, bsb_col).value = bsb
            if chsb_col:
                template_ws.cell(r, chsb_col).value = chsb
            filled += 1
        else:
            skipped += 1
    return {"filled": filled, "skipped": skipped}

def process(template_path: str, data_paths: List[str], teacher_fullname: str) -> Tuple[Workbook, Dict]:
    template_wb = load_workbook(template_path)
    teacher_short = shorten_teacher_name(teacher_fullname)

    # Parse all data files
    class_order: List[str] = []
    data_map: Dict[str, Dict[str, Tuple[Optional[float], Optional[float]]]] = {}
    for p in data_paths:
        wb = load_workbook(p)
        cls, scores = extract_students_scores(wb)
        cls = cls or wb.sheetnames[0]
        class_order.append(cls)
        data_map[cls] = scores

    report = {"teacher_short": teacher_short, "classes": []}

    # Rename template sheets in the same order (up to available sheets)
    tpl_sheets = template_wb.sheetnames[:]
    for i, cls in enumerate(class_order):
        if i >= len(tpl_sheets):
            break
        old = tpl_sheets[i]
        try:
            template_wb[old].title = cls
        except Exception:
            pass

    for cls in class_order:
        if cls not in template_wb.sheetnames:
            continue
        ws = template_wb[cls]
        # teacher
        _set_adjacent(ws, _find_teacher_cells(ws), teacher_short)
        # class
        _set_adjacent(ws, _find_class_cells(ws), cls)
        # fill scores
        stats = _fill_sheet_from_scores(ws, data_map.get(cls, {}))
        report["classes"].append({"class": cls, **stats})

    return template_wb, report
